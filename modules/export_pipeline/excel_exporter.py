"""
Excel Export Module for 1PWR Grid Platform
Generates comprehensive Excel reports with network data, calculations, and visualizations
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart
from datetime import datetime
from typing import Dict, List, Any, Optional
import os


class ExcelExporter:
    """Export network data to formatted Excel reports"""
    
    def __init__(self):
        self.wb = None
        self.status_colors = {
            'as_designed': 'C6EFCE',  # Light green
            'as_built': 'BDD7EE',     # Light blue
            'planned': 'FFE699',       # Light yellow
            'pending': 'FFCC99',       # Light orange
            'error': 'FFC7CE',         # Light red
            'default': 'D9D9D9'        # Light grey
        }
    
    def export_network_report(
        self,
        network_data: Dict[str, Any],
        voltage_results: Optional[Dict[str, Any]] = None,
        validation_results: Optional[Dict[str, Any]] = None,
        site_name: str = "Unknown",
        output_path: str = None
    ) -> str:
        """
        Generate comprehensive network report
        
        Args:
            network_data: Network data dictionary with poles, conductors, etc.
            voltage_results: Voltage calculation results
            validation_results: Network validation results
            site_name: Name of the site
            output_path: Output file path (optional)
            
        Returns:
            Path to generated Excel file
        """
        # Create workbook
        self.wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Add sheets
        self._create_summary_sheet(site_name, network_data, voltage_results, validation_results)
        self._create_poles_sheet(network_data.get('poles', []))
        self._create_conductors_sheet(network_data.get('conductors', []))
        self._create_connections_sheet(network_data.get('connections', []))
        
        if voltage_results:
            self._create_voltage_sheet(voltage_results)
        
        if validation_results:
            self._create_validation_sheet(validation_results)
        
        self._create_statistics_sheet(network_data)
        
        # Generate output path if not provided
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"network_report_{site_name}_{timestamp}.xlsx"
        
        # Save workbook
        self.wb.save(output_path)
        return output_path
    
    def _create_summary_sheet(self, site_name: str, network_data: Dict, 
                             voltage_results: Optional[Dict], 
                             validation_results: Optional[Dict]):
        """Create summary sheet with key metrics"""
        ws = self.wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = f"Network Report - {site_name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Key Metrics
        row = 4
        ws[f'A{row}'] = "KEY METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        metrics = [
            ("Total Poles", len(network_data.get('poles', []))),
            ("Total Conductors", len(network_data.get('conductors', []))),
            ("Total Connections", len(network_data.get('connections', []))),
            ("Total Transformers", len(network_data.get('transformers', []))),
        ]
        
        if voltage_results:
            metrics.extend([
                ("Max Voltage Drop (%)", f"{voltage_results.get('max_drop', 0):.2f}"),
                ("Voltage Violations", len(voltage_results.get('violations', []))),
            ])
        
        if validation_results:
            metrics.extend([
                ("Validation Errors", len(validation_results.get('errors', []))),
                ("Validation Warnings", len(validation_results.get('warnings', []))),
            ])
        
        for metric_name, metric_value in metrics:
            row += 1
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = metric_value
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
        
        # Status Distribution
        row += 3
        ws[f'A{row}'] = "STATUS DISTRIBUTION"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        status_counts = {}
        for pole in network_data.get('poles', []):
            status = pole.get('status', 'unknown')
            status_counts[status] = status_counts.get(status, 0) + 1
        
        row += 1
        ws[f'A{row}'] = "Status"
        ws[f'B{row}'] = "Count"
        ws[f'C{row}'] = "Percentage"
        
        total_poles = len(network_data.get('poles', []))
        for status, count in status_counts.items():
            row += 1
            ws[f'A{row}'] = status
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{(count/total_poles*100):.1f}%" if total_poles > 0 else "0%"
            
            # Apply status color
            fill_color = self.status_colors.get(status, self.status_colors['default'])
            ws[f'A{row}'].fill = PatternFill(start_color=fill_color, 
                                            end_color=fill_color, 
                                            fill_type='solid')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_poles_sheet(self, poles: List[Dict]):
        """Create detailed poles sheet"""
        ws = self.wb.create_sheet("Poles")
        
        if not poles:
            ws['A1'] = "No pole data available"
            return
        
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(poles)
        
        # Write headers
        headers = ['ID', 'Type', 'Status', 'Latitude', 'Longitude', 'Validated']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Write data
        for row_idx, pole in enumerate(poles, 2):
            ws.cell(row=row_idx, column=1, value=pole.get('id', ''))
            ws.cell(row=row_idx, column=2, value=pole.get('type', ''))
            ws.cell(row=row_idx, column=3, value=pole.get('status', ''))
            ws.cell(row=row_idx, column=4, value=pole.get('lat', ''))
            ws.cell(row=row_idx, column=5, value=pole.get('lng', ''))
            ws.cell(row=row_idx, column=6, value='Yes' if pole.get('validated') else 'No')
            
            # Apply status color
            status = pole.get('status', 'default')
            fill_color = self.status_colors.get(status, self.status_colors['default'])
            ws.cell(row=row_idx, column=3).fill = PatternFill(
                start_color=fill_color, 
                end_color=fill_color, 
                fill_type='solid'
            )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_conductors_sheet(self, conductors: List[Dict]):
        """Create detailed conductors sheet"""
        ws = self.wb.create_sheet("Conductors")
        
        if not conductors:
            ws['A1'] = "No conductor data available"
            return
        
        # Write headers
        headers = ['ID', 'From Pole', 'To Pole', 'Type', 'Length (m)', 'Conductor Type']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Write data
        for row_idx, conductor in enumerate(conductors, 2):
            ws.cell(row=row_idx, column=1, value=conductor.get('id', ''))
            ws.cell(row=row_idx, column=2, value=conductor.get('from', ''))
            ws.cell(row=row_idx, column=3, value=conductor.get('to', ''))
            ws.cell(row=row_idx, column=4, value=conductor.get('type', ''))
            ws.cell(row=row_idx, column=5, value=conductor.get('length', 0))
            ws.cell(row=row_idx, column=6, value=conductor.get('conductor_type', ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_connections_sheet(self, connections: List[Dict]):
        """Create detailed connections sheet"""
        ws = self.wb.create_sheet("Connections")
        
        if not connections:
            ws['A1'] = "No connection data available"
            return
        
        # Write headers
        headers = ['ID', 'Type', 'Status', 'Latitude', 'Longitude', 'Connected Pole']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Write data
        for row_idx, conn in enumerate(connections, 2):
            ws.cell(row=row_idx, column=1, value=conn.get('id', ''))
            ws.cell(row=row_idx, column=2, value='Customer Connection')
            ws.cell(row=row_idx, column=3, value=conn.get('status', ''))
            ws.cell(row=row_idx, column=4, value=conn.get('lat', ''))
            ws.cell(row=row_idx, column=5, value=conn.get('lng', ''))
            ws.cell(row=row_idx, column=6, value=conn.get('pole_id', ''))
            
            # Apply status color
            status = conn.get('status', 'default')
            fill_color = self.status_colors.get(status, self.status_colors['default'])
            ws.cell(row=row_idx, column=3).fill = PatternFill(
                start_color=fill_color, 
                end_color=fill_color, 
                fill_type='solid'
            )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_voltage_sheet(self, voltage_results: Dict):
        """Create voltage analysis sheet"""
        ws = self.wb.create_sheet("Voltage Analysis")
        
        ws['A1'] = "VOLTAGE DROP ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Summary metrics
        ws['A3'] = "Maximum Voltage Drop (%)"
        ws['B3'] = f"{voltage_results.get('max_drop', 0):.2f}"
        
        ws['A4'] = "Total Violations"
        ws['B4'] = len(voltage_results.get('violations', []))
        
        # Violations table
        violations = voltage_results.get('violations', [])
        if violations:
            ws['A6'] = "VOLTAGE VIOLATIONS"
            ws['A6'].font = Font(bold=True, size=12)
            
            headers = ['Node ID', 'Voltage (V)', 'Drop (%)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            for row_idx, violation in enumerate(violations, 8):
                ws.cell(row=row_idx, column=1, value=violation.get('node', ''))
                ws.cell(row=row_idx, column=2, value=f"{violation.get('voltage', 0):.0f}")
                ws.cell(row=row_idx, column=3, value=f"{violation.get('drop_percent', 0):.2f}")
        
        # Node voltages
        voltages = voltage_results.get('voltages', {})
        if voltages:
            start_row = len(violations) + 10 if violations else 6
            ws[f'A{start_row}'] = "NODE VOLTAGES"
            ws[f'A{start_row}'].font = Font(bold=True, size=12)
            
            headers = ['Node ID', 'Voltage (V)', 'Drop (%)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row+1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            source_voltage = 11000  # Default, should be passed in
            for idx, (node_id, voltage) in enumerate(voltages.items(), start_row+2):
                drop_percent = ((source_voltage - voltage) / source_voltage) * 100
                ws.cell(row=idx, column=1, value=node_id)
                ws.cell(row=idx, column=2, value=f"{voltage:.0f}")
                ws.cell(row=idx, column=3, value=f"{drop_percent:.2f}")
                
                # Color code based on drop percentage
                if drop_percent > 7:
                    fill_color = 'FFC7CE'  # Red
                elif drop_percent > 5:
                    fill_color = 'FFCC99'  # Orange
                elif drop_percent > 3:
                    fill_color = 'FFE699'  # Yellow
                else:
                    fill_color = 'C6EFCE'  # Green
                
                ws.cell(row=idx, column=3).fill = PatternFill(
                    start_color=fill_color, 
                    end_color=fill_color, 
                    fill_type='solid'
                )
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_validation_sheet(self, validation_results: Dict):
        """Create validation results sheet"""
        ws = self.wb.create_sheet("Validation")
        
        ws['A1'] = "NETWORK VALIDATION RESULTS"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = f"Validation Status: {'PASSED' if validation_results.get('valid') else 'FAILED'}"
        ws['A3'].font = Font(bold=True)
        
        # Errors
        errors = validation_results.get('errors', [])
        if errors:
            ws['A5'] = f"ERRORS ({len(errors)})"
            ws['A5'].font = Font(bold=True, size=12, color='FF0000')
            
            for idx, error in enumerate(errors, 6):
                ws[f'A{idx}'] = f"• {error}"
                ws[f'A{idx}'].font = Font(color='FF0000')
        
        # Warnings
        warnings = validation_results.get('warnings', [])
        if warnings:
            start_row = len(errors) + 7 if errors else 5
            ws[f'A{start_row}'] = f"WARNINGS ({len(warnings)})"
            ws[f'A{start_row}'].font = Font(bold=True, size=12, color='FF9900')
            
            for idx, warning in enumerate(warnings, start_row+1):
                ws[f'A{idx}'] = f"• {warning}"
                ws[f'A{idx}'].font = Font(color='FF9900')
        
        # Adjust column width
        ws.column_dimensions['A'].width = 80
    
    def _create_statistics_sheet(self, network_data: Dict):
        """Create statistics and charts sheet"""
        ws = self.wb.create_sheet("Statistics")
        
        ws['A1'] = "NETWORK STATISTICS"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Component counts
        ws['A3'] = "Component Type"
        ws['B3'] = "Count"
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)
        
        components = [
            ("MV Poles", sum(1 for p in network_data.get('poles', []) if p.get('type') == 'MV')),
            ("LV Poles", sum(1 for p in network_data.get('poles', []) if p.get('type') == 'LV')),
            ("Backbone Conductors", sum(1 for c in network_data.get('conductors', []) if c.get('type') == 'backbone')),
            ("Distribution Conductors", sum(1 for c in network_data.get('conductors', []) if c.get('type') == 'distribution')),
            ("Connections", len(network_data.get('connections', []))),
            ("Transformers", len(network_data.get('transformers', [])))
        ]
        
        for idx, (comp_type, count) in enumerate(components, 4):
            ws[f'A{idx}'] = comp_type
            ws[f'B{idx}'] = count
        
        # Create bar chart
        chart = BarChart()
        chart.title = "Network Components"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Component Type"
        
        data = Reference(ws, min_col=2, min_row=3, max_row=3+len(components))
        categories = Reference(ws, min_col=1, min_row=4, max_row=3+len(components))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        ws.add_chart(chart, "D3")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def export_field_report(
        self,
        site_name: str,
        work_completed: List[Dict],
        pending_work: List[Dict],
        issues: List[Dict],
        output_path: str = None
    ) -> str:
        """
        Generate field team report
        
        Args:
            site_name: Name of the site
            work_completed: List of completed work items
            pending_work: List of pending work items
            issues: List of issues/observations
            output_path: Output file path (optional)
            
        Returns:
            Path to generated Excel file
        """
        self.wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Create field report sheets
        self._create_field_summary(site_name, work_completed, pending_work, issues)
        self._create_work_completed_sheet(work_completed)
        self._create_pending_work_sheet(pending_work)
        self._create_issues_sheet(issues)
        
        # Generate output path if not provided
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"field_report_{site_name}_{timestamp}.xlsx"
        
        # Save workbook
        self.wb.save(output_path)
        return output_path
    
    def _create_field_summary(self, site_name: str, completed: List, pending: List, issues: List):
        """Create field report summary"""
        ws = self.wb.create_sheet("Field Summary")
        
        ws['A1'] = f"Field Report - {site_name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Summary stats
        ws['A4'] = "SUMMARY"
        ws['A4'].font = Font(bold=True, size=12)
        
        stats = [
            ("Work Items Completed", len(completed)),
            ("Work Items Pending", len(pending)),
            ("Issues Reported", len(issues)),
            ("Critical Issues", sum(1 for i in issues if i.get('severity') == 'critical')),
        ]
        
        for idx, (label, value) in enumerate(stats, 5):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _create_work_completed_sheet(self, work_completed: List[Dict]):
        """Create completed work sheet"""
        ws = self.wb.create_sheet("Completed Work")
        
        if not work_completed:
            ws['A1'] = "No completed work items"
            return
        
        headers = ['Item ID', 'Description', 'Component Type', 'Date Completed', 'Team']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        for row_idx, item in enumerate(work_completed, 2):
            ws.cell(row=row_idx, column=1, value=item.get('id', ''))
            ws.cell(row=row_idx, column=2, value=item.get('description', ''))
            ws.cell(row=row_idx, column=3, value=item.get('component_type', ''))
            ws.cell(row=row_idx, column=4, value=item.get('date_completed', ''))
            ws.cell(row=row_idx, column=5, value=item.get('team', ''))
    
    def _create_pending_work_sheet(self, pending_work: List[Dict]):
        """Create pending work sheet"""
        ws = self.wb.create_sheet("Pending Work")
        
        if not pending_work:
            ws['A1'] = "No pending work items"
            return
        
        headers = ['Item ID', 'Description', 'Component Type', 'Priority', 'Assigned Team']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        for row_idx, item in enumerate(pending_work, 2):
            ws.cell(row=row_idx, column=1, value=item.get('id', ''))
            ws.cell(row=row_idx, column=2, value=item.get('description', ''))
            ws.cell(row=row_idx, column=3, value=item.get('component_type', ''))
            ws.cell(row=row_idx, column=4, value=item.get('priority', ''))
            ws.cell(row=row_idx, column=5, value=item.get('assigned_team', ''))
    
    def _create_issues_sheet(self, issues: List[Dict]):
        """Create issues sheet"""
        ws = self.wb.create_sheet("Issues")
        
        if not issues:
            ws['A1'] = "No issues reported"
            return
        
        headers = ['Issue ID', 'Description', 'Severity', 'Component', 'Date Reported']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        for row_idx, issue in enumerate(issues, 2):
            ws.cell(row=row_idx, column=1, value=issue.get('id', ''))
            ws.cell(row=row_idx, column=2, value=issue.get('description', ''))
            ws.cell(row=row_idx, column=3, value=issue.get('severity', ''))
            ws.cell(row=row_idx, column=4, value=issue.get('component', ''))
            ws.cell(row=row_idx, column=5, value=issue.get('date_reported', ''))
            
            # Color code by severity
            severity = issue.get('severity', '').lower()
            if severity == 'critical':
                fill_color = 'FF0000'
            elif severity == 'high':
                fill_color = 'FF9900'
            elif severity == 'medium':
                fill_color = 'FFCC00'
            else:
                fill_color = 'C6EFCE'
            
            ws.cell(row=row_idx, column=3).fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type='solid'
            )
